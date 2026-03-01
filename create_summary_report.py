#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
교수님 공유용 초안 보고서 생성 스크립트
- 변수 정의 (Input/Outcome)
- 샘플 데이터 기반 모델 성능 결과
"""

import json
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def create_summary_report(output_path=None):
    if output_path is None:
        output_path = os.path.join(SCRIPT_DIR, 'results', '분석_결과_초안_보고서.xlsx')

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()

    # ---- Styles ----
    header_font = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    subheader_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    subheader_font = Font(name='맑은 고딕', bold=True, size=10)
    normal_font = Font(name='맑은 고딕', size=10)
    title_font = Font(name='맑은 고딕', bold=True, size=14, color='2F5496')
    subtitle_font = Font(name='맑은 고딕', bold=True, size=12, color='2F5496')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    def style_rows(ws, start_row, end_row, max_col):
        alt_fill = PatternFill(start_color='F2F7FC', end_color='F2F7FC', fill_type='solid')
        for r in range(start_row, end_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if (r - start_row) % 2 == 1:
                    cell.fill = alt_fill

    # =========================================================================
    # Sheet 1: 연구 개요
    # =========================================================================
    ws1 = wb.active
    ws1.title = '연구 개요'

    overview = [
        ('연구 제목', '담석증 환자에서의 당뇨병 발생 예측 모델 개발'),
        ('연구 목적', '담석증(Cholelithiasis) 진단 환자 코호트에서 머신러닝 기반 당뇨병 발생 예측 모델을 개발하고,\n'
                   '모델 간 성능을 비교하여 최적 모델을 선정'),
        ('데이터 소스', '국민건강보험공단 건강검진 코호트 (더미 데이터로 검증)'),
        ('분석 대상', '담석증 진단 이력이 있는 성인 (만 30세 이상)'),
        ('', ''),
        ('분석 방법', ''),
        ('  전처리', '- 연속형 변수: 결측치 중앙값 대체 + StandardScaler 정규화\n'
                   '- 범주형 변수: 결측치 최빈값 대체\n'
                   '- Missing Indicator: 결측률 5% 이상 변수에 결측 여부 지시 변수 추가 (LDL, Proteinuria)'),
        ('  데이터 분할', 'Train : Validation : Test = 70% : 10% : 20% (Stratified)'),
        ('  모델 학습', 'GridSearchCV (5-fold CV, scoring=AUROC)'),
        ('  성능 평가', 'AUROC, AUPRC, Accuracy, Sensitivity, Specificity, PPV, NPV, F1 Score\n'
                     '+ Bootstrap 95% CI'),
        ('  모델 해석', 'SHAP (SHapley Additive exPlanations) analysis'),
        ('', ''),
        ('Outcome', ''),
        ('  outA', '당뇨병 발생 (모든 유형)'),
        ('  out2', '2형 당뇨병 발생'),
        ('', ''),
        ('사용 모델 (6개)', ''),
        ('  1. Logistic Regression', '기준 모델 (선형 모델)'),
        ('  2. Decision Tree', '해석 가능한 트리 모델'),
        ('  3. Random Forest', '앙상블 (Bagging)'),
        ('  4. XGBoost', '앙상블 (Gradient Boosting)'),
        ('  5. LightGBM', '앙상블 (Gradient Boosting, 경량)'),
        ('  6. ANN (MLP)', '인공신경망 (다층 퍼셉트론)'),
    ]

    ws1.cell(row=1, column=1, value='담석증 환자 당뇨병 예측 모델 - 분석 결과 초안').font = title_font
    ws1.merge_cells('A1:B1')

    for i, (key, val) in enumerate(overview, start=3):
        cell_a = ws1.cell(row=i, column=1, value=key)
        cell_b = ws1.cell(row=i, column=2, value=val)
        cell_a.font = Font(name='맑은 고딕', bold=True, size=10)
        cell_b.font = normal_font
        cell_b.alignment = Alignment(wrap_text=True, vertical='top')

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 80

    # =========================================================================
    # Sheet 2: 변수 정의
    # =========================================================================
    ws2 = wb.create_sheet('변수 정의')

    ws2.cell(row=1, column=1, value='Input / Outcome 변수 정의').font = subtitle_font
    ws2.merge_cells('A1:F1')

    headers = ['구분', '변수명 (원본)', '표시명', '데이터 타입', '설명', '비고']
    row = 3
    for c, h in enumerate(headers, 1):
        ws2.cell(row=row, column=c, value=h)
    style_header(ws2, row, len(headers))

    variables = [
        # Continuous Input
        ('Input (연속형)', 'age', 'Age (years)', '연속형', '연령', ''),
        ('Input (연속형)', 'BMI', 'BMI (kg/m²)', '연속형', '체질량지수', '결측 ~3%'),
        ('Input (연속형)', 'SBP', 'SBP (mmHg)', '연속형', '수축기 혈압', '결측 ~2%'),
        ('Input (연속형)', 'DBP', 'DBP (mmHg)', '연속형', '이완기 혈압', '결측 ~2%'),
        ('Input (연속형)', 'FBS', 'Glucose (mg/dL)', '연속형', '공복혈당', '결측 ~3%'),
        ('Input (연속형)', 'TOT_CHOL', 'Total cholesterol (mg/dL)', '연속형', '총 콜레스테롤', '결측 ~3%'),
        ('Input (연속형)', 'WAIST', 'Waist (cm)', '연속형', '허리둘레', '결측 ~4%'),
        ('Input (연속형)', 'TG', 'Triglyceride (mg/dL)', '연속형', '중성지방', '결측 ~4%'),
        ('Input (연속형)', 'HDL_CHOL', 'HDL cholesterol (mg/dL)', '연속형', 'HDL 콜레스테롤', '결측 ~4%'),
        ('Input (연속형)', 'LDL_CHOL', 'LDL cholesterol (mg/dL)', '연속형', 'LDL 콜레스테롤', '결측 ~30%, Missing Indicator 추가'),
        ('Input (연속형)', 'Creatinine', 'Creatinine (mg/dL)', '연속형', '크레아티닌', '결측 ~3%'),
        # Categorical Input
        ('Input (범주형)', 'diag', 'Cholelithiasis', '이진 (0/1)', '담석증 진단 유무', '0=No, 1=Yes'),
        ('Input (범주형)', 'act', 'Cholecystectomy', '이진 (0/1)', '담낭절제술 수술 유무', '0=No, 1=Yes'),
        ('Input (범주형)', 'gender', 'Sex', '이진 (0/1)', '성별', '0=Male, 1=Female'),
        ('Input (범주형)', 'smoking', 'Smoking', '범주 (0/1/2)', '흡연 상태', '0=Never, 1=Former, 2=Current'),
        ('Input (범주형)', 'drink', 'Alcohol', '이진 (0/1)', '주 2일 이상 음주', '0=No, 1=Yes'),
        ('Input (범주형)', 'training', 'Training', '이진 (0/1)', '주 3일 이상 운동', '0=No, 1=Yes'),
        ('Input (범주형)', 'proteinUria', 'Proteinuria', '범주 (0/1/2)', '단백뇨', '0=Normal, 1=Trace/+1, 2=≥+2\nMissing Indicator 추가'),
        ('Input (범주형)', 'co_HLD', 'Dyslipidemia', '이진 (0/1)', '이상지질혈증 동반', '0=No, 1=Yes'),
        ('Input (범주형)', 'co_HTN', 'Hypertension', '이진 (0/1)', '고혈압 동반', '0=No, 1=Yes'),
        ('Input (범주형)', 'co_fattyLiver', 'Fatty liver', '이진 (0/1)', '지방간 동반', '0=No, 1=Yes'),
        ('Input (범주형)', 'co_Impaird', 'Impaired fasting glucose', '이진 (0/1)', '공복혈당장애 동반', '0=No, 1=Yes'),
        ('Input (범주형)', 'metS', 'Metabolic syndrome', '이진 (0/1)', '대사증후군 유무', '0=No, 1=Yes'),
        # Missing Indicator
        ('Input (결측 지시)', 'LDL_CHOL_missing', 'LDL cholesterol_missing', '이진 (0/1)', 'LDL cholesterol 결측 여부', '자동 생성'),
        ('Input (결측 지시)', 'proteinUria_missing', 'Proteinuria_missing', '이진 (0/1)', 'Proteinuria 결측 여부', '자동 생성'),
        # Outcome
        ('Outcome', 'outA', 'Diabetes incidence', '이진 (0/1)', '당뇨병 발생 (모든 유형)', '주요 결과 변수'),
        ('Outcome', 'out2', 'Type 2 Diabetes incidence', '이진 (0/1)', '2형 당뇨병 발생', '보조 결과 변수'),
    ]

    start_row = row + 1
    for i, v in enumerate(variables):
        for c, val in enumerate(v, 1):
            ws2.cell(row=start_row + i, column=c, value=val)
    style_rows(ws2, start_row, start_row + len(variables) - 1, len(headers))

    # Outcome rows highlighting
    outcome_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    for i, v in enumerate(variables):
        if v[0] == 'Outcome':
            for c in range(1, len(headers) + 1):
                ws2.cell(row=start_row + i, column=c).fill = outcome_fill

    for c, w in enumerate([18, 22, 30, 16, 30, 35], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # =========================================================================
    # Sheet 3: 모델 성능 비교 (outA)
    # =========================================================================
    targets = ['outA', 'out2']
    target_labels = {'outA': '당뇨병 발생 (outA)', 'out2': '2형 당뇨병 발생 (out2)'}
    model_names_display = {
        'logistic_regression': 'Logistic Regression',
        'decision_tree': 'Decision Tree',
        'random_forest': 'Random Forest',
        'xgboost': 'XGBoost',
        'lightgbm': 'LightGBM',
        'ann': 'ANN (MLP)'
    }
    model_order = ['logistic_regression', 'decision_tree', 'random_forest',
                   'xgboost', 'lightgbm', 'ann']

    for target in targets:
        ws = wb.create_sheet(f'성능 비교 ({target})')
        ws.cell(row=1, column=1,
                value=f'모델 성능 비교 - {target_labels[target]}').font = subtitle_font
        ws.merge_cells('A1:I1')
        ws.cell(row=2, column=1,
                value='* 더미 데이터 (N=10,000) 기반 결과 - 실제 데이터 적용 시 성능 변동 예상').font = Font(
            name='맑은 고딕', size=9, italic=True, color='888888')

        perf_headers = ['Model', 'AUROC', 'AUPRC', 'Accuracy', 'Sensitivity',
                        'Specificity', 'PPV', 'NPV', 'F1 Score']
        row = 4
        for c, h in enumerate(perf_headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header(ws, row, len(perf_headers))

        data_row = row + 1
        best_auroc = 0
        best_row = data_row

        for i, model in enumerate(model_order):
            metrics_path = os.path.join(SCRIPT_DIR, 'results', target, model, 'metrics.json')
            if not os.path.exists(metrics_path):
                continue
            with open(metrics_path) as f:
                m = json.load(f)

            r = data_row + i
            ws.cell(row=r, column=1, value=model_names_display[model])
            ws.cell(row=r, column=2, value=round(m['auroc'], 4))
            ws.cell(row=r, column=3, value=round(m['auprc'], 4))
            ws.cell(row=r, column=4, value=round(m['accuracy'], 4))
            ws.cell(row=r, column=5, value=round(m['sensitivity'], 4))
            ws.cell(row=r, column=6, value=round(m['specificity'], 4))
            ws.cell(row=r, column=7, value=round(m['ppv'], 4))
            ws.cell(row=r, column=8, value=round(m['npv'], 4))
            ws.cell(row=r, column=9, value=round(m['f1'], 4))

            if m['auroc'] > best_auroc:
                best_auroc = m['auroc']
                best_row = r

        end_data_row = data_row + len(model_order) - 1
        style_rows(ws, data_row, end_data_row, len(perf_headers))

        # Highlight best model
        best_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        for c in range(1, len(perf_headers) + 1):
            ws.cell(row=best_row, column=c).fill = best_fill

        ws.cell(row=end_data_row + 2, column=1,
                value='* 초록색 음영 = AUROC 기준 최고 성능 모델').font = Font(
            name='맑은 고딕', size=9, italic=True, color='888888')

        for c, w in enumerate([22, 12, 12, 12, 12, 12, 12, 12, 12], 1):
            ws.column_dimensions[get_column_letter(c)].width = w

    # =========================================================================
    # Sheet 5: 성능 요약 (두 타겟 비교)
    # =========================================================================
    ws5 = wb.create_sheet('성능 요약 비교')
    ws5.cell(row=1, column=1, value='Outcome별 모델 성능 비교 요약 (AUROC)').font = subtitle_font
    ws5.merge_cells('A1:D1')

    sum_headers = ['Model', 'outA (당뇨병)', 'out2 (2형 당뇨병)', '비고']
    row = 3
    for c, h in enumerate(sum_headers, 1):
        ws5.cell(row=row, column=c, value=h)
    style_header(ws5, row, len(sum_headers))

    data_row = row + 1
    for i, model in enumerate(model_order):
        r = data_row + i
        ws5.cell(row=r, column=1, value=model_names_display[model])

        for j, target in enumerate(targets, 2):
            metrics_path = os.path.join(SCRIPT_DIR, 'results', target, model, 'metrics.json')
            if os.path.exists(metrics_path):
                with open(metrics_path) as f:
                    m = json.load(f)
                ws5.cell(row=r, column=j, value=round(m['auroc'], 4))

        ws5.cell(row=r, column=4, value='')

    style_rows(ws5, data_row, data_row + len(model_order) - 1, len(sum_headers))

    ws5.cell(row=data_row + len(model_order) + 1, column=1,
             value='* 더미 데이터 기반 결과입니다. 실제 NHIS 데이터 적용 시 성능이 달라질 수 있습니다.').font = Font(
        name='맑은 고딕', size=9, italic=True, color='888888')

    for c, w in enumerate([22, 18, 22, 30], 1):
        ws5.column_dimensions[get_column_letter(c)].width = w

    # =========================================================================
    # Sheet 6: 생성 산출물 목록
    # =========================================================================
    ws6 = wb.create_sheet('산출물 목록')
    ws6.cell(row=1, column=1, value='분석 산출물 목록').font = subtitle_font
    ws6.merge_cells('A1:C1')

    out_headers = ['산출물', '경로', '설명']
    row = 3
    for c, h in enumerate(out_headers, 1):
        ws6.cell(row=row, column=c, value=h)
    style_header(ws6, row, len(out_headers))

    outputs = [
        ('Table 1 (Train vs Test)', 'results/{target}/tables/table1_train_test.xlsx', 'Train/Test 데이터 균형 확인용'),
        ('Table 1 (Outcome별)', 'results/{target}/tables/table1_by_{target}.xlsx', 'Outcome별 Baseline characteristics'),
        ('모델 성능 테이블', 'results/{target}/tables/model_performance.xlsx', '전체 모델 성능 비교 (Bootstrap 95% CI 포함)'),
        ('ROC Curve (개별)', 'results/{target}/{model}/roc_curve.png', '각 모델별 ROC curve'),
        ('PR Curve (개별)', 'results/{target}/{model}/pr_curve.png', '각 모델별 Precision-Recall curve'),
        ('Confusion Matrix', 'results/{target}/{model}/confusion_matrix.png', '각 모델별 혼동 행렬'),
        ('Calibration Curve', 'results/{target}/{model}/calibration_curve.png', '각 모델별 보정 곡선'),
        ('SHAP Summary Plot', 'results/{target}/{model}/shap_summary.png', '각 모델별 SHAP 요약 플롯'),
        ('SHAP Bar Plot', 'results/{target}/{model}/shap_bar.png', '각 모델별 SHAP 중요도 막대 그래프'),
        ('ROC 비교 (전체 모델)', 'results/{target}/comparison/comparison_roc.png', '모든 모델 ROC curve 비교'),
        ('PR 비교 (전체 모델)', 'results/{target}/comparison/comparison_pr.png', '모든 모델 PR curve 비교'),
        ('Calibration 비교', 'results/{target}/comparison/comparison_calibration.png', '모든 모델 보정 곡선 비교'),
        ('SHAP 비교', 'results/{target}/comparison/comparison_shap.png', '모든 모델 SHAP 중요도 비교'),
        ('Combined Figure', 'results/{target}/comparison/comparison_combined.png', 'ROC+PR+Calibration 결합 Figure'),
    ]

    data_row = row + 1
    for i, (name, path, desc) in enumerate(outputs):
        ws6.cell(row=data_row + i, column=1, value=name)
        ws6.cell(row=data_row + i, column=2, value=path)
        ws6.cell(row=data_row + i, column=3, value=desc)
    style_rows(ws6, data_row, data_row + len(outputs) - 1, len(out_headers))

    ws6.cell(row=data_row + len(outputs) + 1, column=1,
             value='* {target} = outA 또는 out2, {model} = 모델명 (logistic_regression, random_forest 등)').font = Font(
        name='맑은 고딕', size=9, italic=True, color='888888')
    ws6.cell(row=data_row + len(outputs) + 2, column=1,
             value='* 모든 Figure는 .png, .tiff, .pdf 3가지 형식으로 저장됨').font = Font(
        name='맑은 고딕', size=9, italic=True, color='888888')

    for c, w in enumerate([25, 55, 40], 1):
        ws6.column_dimensions[get_column_letter(c)].width = w

    # =========================================================================
    # Save
    # =========================================================================
    wb.save(output_path)
    print(f"\n✅ 초안 보고서 저장 완료: {output_path}")
    return output_path


if __name__ == '__main__':
    create_summary_report()
